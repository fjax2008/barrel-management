"""
生管部在制品桶号管理系统 - 后端服务
FastAPI + Turso 云数据库（数据永不丢失）
"""
from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import asyncio
import io
import os
from datetime import datetime, timezone, timedelta
from libsql_client import create_client

app = FastAPI(title="生管部在制品桶号管理系统")

# 时区
TZ = timezone(timedelta(hours=7))  # GMT+7 (越南/中国)
def now_str(fmt="%Y-%m-%d %H:%M:%S"):
    return datetime.now(TZ).strftime(fmt)

# ─── Turso 云数据库 ───────────────────────────────────────

TURSO_URL = os.environ.get("TURSO_URL", "https://barrel-db-fjax2008.aws-ap-south-1.turso.io")
TURSO_TOKEN = os.environ.get("TURSO_TOKEN", "eyJhbGciOiJFZERTQSIsInR5cCI6IkpXVCJ9.eyJpYXQiOjE3ODEyNDk3NDQsImlkIjoiMDE5ZWJhYzItMTAwMS03OGYzLWIzNjItNzA1MDlmNGFmYWY1IiwicmlkIjoiODJlNjM0YWItNjc3Ni00MTgyLTkxNjMtN2QzOWFhYzA4OGFiIn0.kzTmT2Y8jU47VBP_XIYq0QHXvNkwnk8jTBDVsAEHCxiV0OYKGGf6iuypDiM_-H9U5KLnSBMFaicrkV6h9TDGDw")


def db(sql: str, params=None):
    """同步执行 SQL 并返回 [dict] 列表"""
    async def _exec():
        client = create_client(url=TURSO_URL, auth_token=TURSO_TOKEN)
        try:
            rs = await client.execute(sql, params or [])
            return [dict(zip(rs.columns, row)) for row in rs.rows]
        finally:
            await client.close()
    try:
        return asyncio.run(_exec())
    except RuntimeError:
        loop = asyncio.get_event_loop()
        if loop.is_running():
            import nest_asyncio
            nest_asyncio.apply()
            return asyncio.run(_exec())
        return loop.run_until_complete(_exec())


def db_batch(stmts: list):
    """批量执行多条 SQL（事务性执行）"""
    async def _exec():
        client = create_client(url=TURSO_URL, auth_token=TURSO_TOKEN)
        try:
            results = await client.batch(stmts)
            # 检查每个结果是否有错误
            for i, rs in enumerate(results):
                if hasattr(rs, 'error') and rs.error:
                    print(f"[DB BATCH ERROR] idx={i}: {rs.error}")
            return results
        finally:
            await client.close()
    try:
        return asyncio.run(_exec())
    except RuntimeError:
        loop = asyncio.get_event_loop()
        if loop.is_running():
            import nest_asyncio
            nest_asyncio.apply()
            return asyncio.run(_exec())
        return loop.run_until_complete(_exec())


# ─── 启动初始化 ───────────────────────────────────────────

def init_db():
    db("""CREATE TABLE IF NOT EXISTS barrel_inventory (
        barrel_no TEXT PRIMARY KEY,
        location TEXT NOT NULL,
        status TEXT NOT NULL DEFAULT 'in',
        update_time TEXT NOT NULL
    )""")
    db("""CREATE TABLE IF NOT EXISTS barrel_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        barrel_no TEXT NOT NULL,
        action TEXT NOT NULL,
        location TEXT NOT NULL,
        destination TEXT NOT NULL DEFAULT '',
        created_at TEXT NOT NULL
    )""")


# ─── Models ───────────────────────────────────────────────

class InboundRequest(BaseModel):
    barrel_no: str
    location: str


class OutboundRequest(BaseModel):
    barrel_no: str
    destination: str = ""


class BatchInboundRequest(BaseModel):
    barrel_nos: list[str]
    location: str


class BatchOutboundRequest(BaseModel):
    barrel_nos: list[str]
    destination: str = ""


class DeleteRequest(BaseModel):
    barrel_nos: list[str]


# ─── 单桶入库 ─────────────────────────────────────────────

@app.post("/api/inbound")
def inbound(req: InboundRequest):
    now = now_str("%Y-%m-%d %H:%M:%S")

    rows = db("SELECT * FROM barrel_inventory WHERE barrel_no = ?", [req.barrel_no])
    if rows and rows[0]["status"] == "in":
        return {"success": False, "msg": f"桶号 {req.barrel_no} 已在 {rows[0]['location']}，请先出库再入库"}

    db_batch([
        ("INSERT OR REPLACE INTO barrel_inventory (barrel_no, location, status, update_time) VALUES (?, ?, 'in', ?)",
         [req.barrel_no, req.location, now]),
        ("INSERT INTO barrel_log (barrel_no, action, location, created_at) VALUES (?, '入库', ?, ?)",
         [req.barrel_no, req.location, now]),
    ])
    return {"success": True, "msg": f"桶号 {req.barrel_no} 已入库 → {req.location}"}


# ─── 批量入库 ─────────────────────────────────────────────

@app.post("/api/inbound/batch")
def inbound_batch(req: BatchInboundRequest):
    now = now_str("%Y-%m-%d %H:%M:%S")
    clean = [bn.strip() for bn in req.barrel_nos if bn.strip()]
    if not clean:
        return {"success": True, "location": req.location, "success_count": 0, "success_list": []}
    # 批量查询已有桶号（一趟往返）
    placeholders = ",".join(["?"] * len(clean))
    existing = db(f"SELECT barrel_no FROM barrel_inventory WHERE barrel_no IN ({placeholders})", clean)
    exists_set = {r["barrel_no"] for r in existing}
    # 批量写入（一趟往返）
    stmts = []
    for bn in clean:
        action = "移动" if bn in exists_set else "入库"
        stmts.append(("INSERT OR REPLACE INTO barrel_inventory (barrel_no, location, status, update_time) VALUES (?, ?, 'in', ?)",
                      [bn, req.location, now]))
        stmts.append(("INSERT INTO barrel_log (barrel_no, action, location, created_at) VALUES (?, ?, ?, ?)",
                      [bn, action, req.location, now]))
    db_batch(stmts)
    return {"success": True, "location": req.location, "success_count": len(clean),
            "success_list": clean}


# ─── 单桶出库 ─────────────────────────────────────────────

@app.post("/api/outbound")
def outbound(req: OutboundRequest):
    now = now_str("%Y-%m-%d %H:%M:%S")

    rows = db("SELECT * FROM barrel_inventory WHERE barrel_no = ? AND status = 'in'", [req.barrel_no])
    if not rows:
        return {"success": False, "msg": f"桶号 {req.barrel_no} 不在库中，无法出库"}

    loc = rows[0]["location"]
    dest = req.destination.strip()
    db_batch([
        ("UPDATE barrel_inventory SET status = 'out', update_time = ? WHERE barrel_no = ?", [now, req.barrel_no]),
        ("INSERT INTO barrel_log (barrel_no, action, location, destination, created_at) VALUES (?, '出库', ?, ?, ?)",
         [req.barrel_no, loc, dest, now]),
    ])
    dest_info = f" → {dest}" if dest else ""
    return {"success": True, "msg": f"桶号 {req.barrel_no} 已从 {loc} 出库{dest_info}", "from_location": loc}


# ─── 批量出库 ─────────────────────────────────────────────

@app.post("/api/outbound/batch")
def outbound_batch(req: BatchOutboundRequest):
    now = now_str("%Y-%m-%d %H:%M:%S")
    success_list, not_found_list = [], []
    dest = req.destination.strip()

    for barrel_no in req.barrel_nos:
        barrel_no = barrel_no.strip()
        if not barrel_no:
            continue
        rows = db("SELECT * FROM barrel_inventory WHERE barrel_no = ? AND status = 'in'", [barrel_no])
        if not rows:
            not_found_list.append(barrel_no)
            continue
        loc = rows[0]["location"]
        db_batch([
            ("UPDATE barrel_inventory SET status = 'out', update_time = ? WHERE barrel_no = ?", [now, barrel_no]),
            ("INSERT INTO barrel_log (barrel_no, action, location, destination, created_at) VALUES (?, '出库', ?, ?, ?)",
             [barrel_no, loc, dest, now]),
        ])
        success_list.append({"barrel_no": barrel_no, "from_location": loc})

    return {"success": True, "success_count": len(success_list), "not_found_count": len(not_found_list),
            "success_list": success_list, "not_found_list": not_found_list}


# ─── 删除桶号 ─────────────────────────────────────────────

@app.delete("/api/barrel/{barrel_no}")
def delete_barrel(barrel_no: str):
    now = now_str("%Y-%m-%d %H:%M:%S")
    rows = db("SELECT * FROM barrel_inventory WHERE barrel_no = ?", [barrel_no])
    if not rows:
        raise HTTPException(404, f"桶号 {barrel_no} 不存在")
    loc = rows[0]["location"]
    status = rows[0]["status"]
    db_batch([
        ("DELETE FROM barrel_inventory WHERE barrel_no = ?", [barrel_no]),
        ("INSERT INTO barrel_log (barrel_no, action, location, destination, created_at) VALUES (?, '删除', ?, '', ?)",
         [barrel_no, loc, now]),
    ])
    return {"success": True, "msg": f"桶号 {barrel_no}（{status} / {loc}）已删除"}


@app.post("/api/barrel/delete/batch")
def delete_barrel_batch(req: DeleteRequest):
    now = now_str("%Y-%m-%d %H:%M:%S")
    deleted, not_found = [], []

    for barrel_no in req.barrel_nos:
        barrel_no = barrel_no.strip()
        if not barrel_no:
            continue
        rows = db("SELECT * FROM barrel_inventory WHERE barrel_no = ?", [barrel_no])
        if not rows:
            not_found.append(barrel_no)
            continue
        loc = rows[0]["location"]
        db_batch([
            ("DELETE FROM barrel_inventory WHERE barrel_no = ?", [barrel_no]),
            ("INSERT INTO barrel_log (barrel_no, action, location, destination, created_at) VALUES (?, '删除', ?, '', ?)",
             [barrel_no, loc, now]),
        ])
        deleted.append(barrel_no)

    return {"success": True, "deleted": deleted, "not_found": not_found,
            "msg": f"已删除 {len(deleted)} 桶" + (f"，{len(not_found)} 桶不存在" if not_found else "")}


# ─── 数据恢复 ─────────────────────────────────────────────

@app.post("/api/import")
def import_data(data: list[dict]):
    """批量导入备份数据"""
    now = now_str("%Y-%m-%d %H:%M:%S")
    imported = 0
    for item in data:
        bn = item.get("barrel_no", "").strip()
        loc = item.get("location", "").strip()
        if not bn or not loc:
            continue
        db_batch([
            ("INSERT OR REPLACE INTO barrel_inventory (barrel_no, location, status, update_time) VALUES (?, ?, 'in', ?)",
             [bn, loc, now]),
            ("INSERT INTO barrel_log (barrel_no, action, location, destination, created_at) VALUES (?, '入库', ?, '', ?)",
             [bn, loc, now]),
        ])
        imported += 1
    return {"success": True, "msg": f"已恢复 {imported} 条记录"}


# ─── 查询 ─────────────────────────────────────────────────

@app.get("/api/inventory")
def get_inventory():
    return db("SELECT * FROM barrel_inventory WHERE status = 'in' ORDER BY location")


@app.get("/api/inventory/{keyword}")
def query_barrel(keyword: str):
    """模糊查询：按桶号或储位搜索"""
    rows = db("SELECT * FROM barrel_inventory WHERE barrel_no LIKE ? ORDER BY status='in' DESC, update_time DESC",
              [f"%{keyword}%"])
    if not rows:
        rows = db("SELECT * FROM barrel_inventory WHERE location LIKE ? AND status = 'in' ORDER BY location",
                  [f"%{keyword}%"])
    return rows


@app.get("/api/inventory/location/{location}")
def query_location(location: str):
    return db("SELECT * FROM barrel_inventory WHERE location = ? AND status = 'in'", [location])


@app.get("/api/logs")
def get_logs(limit: int = 100):
    return db("SELECT * FROM barrel_log ORDER BY id DESC LIMIT ?", [limit])


# ─── Excel 导出 ───────────────────────────────────────────

@app.get("/api/export")
def export_inventory():
    try:
        from urllib.parse import quote
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl 未安装，请执行 pip install openpyxl")

    rows = db("SELECT * FROM barrel_inventory WHERE status = 'in' ORDER BY location")

    wb = Workbook()
    ws = wb.active
    ws.title = "在库清单"

    header_font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    cell_font = Font(name="微软雅黑", size=11)
    cell_align = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:D1")
    ws["A1"] = "生管部在制品桶号管理 - 库存清单"
    ws["A1"].font = Font(name="微软雅黑", bold=True, size=14, color="1677FF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = f"导出时间: {now_str('%Y-%m-%d %H:%M:%S')}　　共 {len(rows)} 桶在库"
    ws["A2"].font = Font(name="微软雅黑", size=10, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["序号", "桶号", "储位", "入库时间"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 1):
        values = [row_idx, row["barrel_no"], row["location"], row["update_time"]]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx + 4, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[4].height = 24

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"桶号库存_{now_str('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


# ─── 静态文件 ─────────────────────────────────────────────

app.mount("/", StaticFiles(directory="static", html=True), name="static")

init_db()
print("✅ Turso 数据库连接成功")
