"""
生管部在制品桶号管理系统 — 后端服务
本地 SQLite + Turso 云端备份
"""
from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import sqlite3
import asyncio
import io
import os
import threading
import time
from datetime import datetime, timezone, timedelta
from contextlib import contextmanager

app = FastAPI(title="生管部在制品桶号管理系统")

# 时区
TZ = timezone(timedelta(hours=7))
def now_str(fmt="%Y-%m-%d %H:%M:%S"):
    return datetime.now(TZ).strftime(fmt)

# ─── 本地 SQLite ──────────────────────────────────────────
DB_PATH = "/opt/barrel-management/barrel.db"

@contextmanager
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()

def init_db():
    with get_db() as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS barrel_inventory (
            barrel_no TEXT PRIMARY KEY,
            location TEXT NOT NULL,
            status TEXT DEFAULT 'in',
            update_time TEXT
        )""")
        conn.execute("""CREATE TABLE IF NOT EXISTS barrel_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            barrel_no TEXT NOT NULL,
            action TEXT NOT NULL,
            location TEXT,
            created_at TEXT
        )""")

# ─── Turso 备份 ───────────────────────────────────────────
TURSO_URL = os.environ.get("TURSO_URL", "https://barrel-db-fjax2008.aws-ap-south-1.turso.io")
TURSO_TOKEN = os.environ.get("TURSO_TOKEN", "eyJhbGciOiJFZERTQSIsInR5cCI6IkpXVCJ9.eyJpYXQiOjE3ODEyNDk3NDQsImlkIjoiMDE5ZWJhYzAtODkzYy03ZjYyLWIyMjUtMDU3ZTdiZGYyYzk0In0.D0MzszPVPj49eKhbGd3xLyYKgLx5M3VRsBpZFqNNffxrYQp7LJCx5JfjSQFgBMP3IOLQ3VoTMRoBikmMCZNcCA")

def turso(sql, params=None):
    from libsql_client import create_client
    async def _exec():
        client = create_client(url=TURSO_URL, auth_token=TURSO_TOKEN)
        try:
            rs = await client.execute(sql, params or [])
            return [dict(zip(rs.columns, row)) for row in rs.rows]
        finally:
            await client.close()
    try:
        return asyncio.run(_exec())
    except RuntimeError:
        import nest_asyncio; nest_asyncio.apply()
        return asyncio.run(_exec())

def backup_to_turso():
    """全量同步到 Turso（幂等）"""
    try:
        with get_db() as conn:
            rows = conn.execute("SELECT * FROM barrel_inventory").fetchall()
        for row in rows:
            turso("INSERT OR REPLACE INTO barrel_inventory (barrel_no, location, status, update_time) VALUES (?,?,?,?)",
                  [row["barrel_no"], row["location"], row["status"], row["update_time"]])
        print(f"[BACKUP] {len(rows)} 条已同步到 Turso")
    except Exception as e:
        print(f"[BACKUP ERROR] {e}")

def backup_loop():
    """每天 3:00 AM 自动备份"""
    while True:
        time.sleep(3600)  # 每小时检查一次
        if datetime.now(TZ).hour == 3 and datetime.now(TZ).minute < 5:
            backup_to_turso()
            time.sleep(300)  # 备份完等5分钟避免重复

# 启动备份线程
threading.Thread(target=backup_loop, daemon=True).start()

# ─── 模型 ─────────────────────────────────────────────────
class InboundRequest(BaseModel):
    barrel_no: str
    location: str

class BatchInboundRequest(BaseModel):
    barrel_nos: list[str]
    location: str

class DeleteRequest(BaseModel):
    barrel_nos: list[str]

# ─── API ──────────────────────────────────────────────────

@app.get("/api/inventory")
def get_inventory():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM barrel_inventory WHERE status='in' ORDER BY location").fetchall()
    return [dict(r) for r in rows]

@app.get("/api/inventory/{keyword}")
def query_barrel(keyword: str):
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM barrel_inventory WHERE barrel_no LIKE ? ORDER BY status='in' DESC, update_time DESC",
                            [f"%{keyword}%"]).fetchall()
        if not rows:
            rows = conn.execute("SELECT * FROM barrel_inventory WHERE location LIKE ? AND status='in' ORDER BY location",
                                [f"%{keyword}%"]).fetchall()
    return [dict(r) for r in rows]

@app.get("/api/inventory/location/{location}")
def query_location(location: str):
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM barrel_inventory WHERE location=? AND status='in'", [location]).fetchall()
    return [dict(r) for r in rows]

@app.post("/api/inbound")
def inbound(req: InboundRequest):
    now = now_str()
    with get_db() as conn:
        row = conn.execute("SELECT * FROM barrel_inventory WHERE barrel_no=?", [req.barrel_no]).fetchone()
        if row and row["status"] == "in":
            return {"success": False, "msg": f"桶号 {req.barrel_no} 已在 {row['location']}，请先出库再入库"}
        conn.execute("INSERT OR REPLACE INTO barrel_inventory (barrel_no, location, status, update_time) VALUES (?,?,'in',?)",
                     [req.barrel_no, req.location, now])
        conn.execute("INSERT INTO barrel_log (barrel_no, action, location, created_at) VALUES (?,?,?,?)",
                     [req.barrel_no, "入库", req.location, now])
    return {"success": True, "msg": f"桶号 {req.barrel_no} 已入库 → {req.location}"}

@app.post("/api/inbound/batch")
def inbound_batch(req: BatchInboundRequest):
    now = now_str()
    clean = [bn.strip() for bn in req.barrel_nos if bn.strip()]
    if not clean:
        return {"success": True, "location": req.location, "success_count": 0, "success_list": []}
    with get_db() as conn:
        placeholders = ",".join(["?"] * len(clean))
        existing_rows = conn.execute(f"SELECT barrel_no FROM barrel_inventory WHERE barrel_no IN ({placeholders})", clean).fetchall()
        exists_set = {r["barrel_no"] for r in existing_rows}
        for bn in clean:
            action = "移动" if bn in exists_set else "入库"
            conn.execute("INSERT OR REPLACE INTO barrel_inventory (barrel_no, location, status, update_time) VALUES (?,?,'in',?)",
                         [bn, req.location, now])
            conn.execute("INSERT INTO barrel_log (barrel_no, action, location, created_at) VALUES (?,?,?,?)",
                         [bn, action, req.location, now])
    return {"success": True, "location": req.location, "success_count": len(clean), "success_list": clean}

@app.delete("/api/barrel/{barrel_no}")
def delete_barrel(barrel_no: str):
    now = now_str()
    with get_db() as conn:
        row = conn.execute("SELECT * FROM barrel_inventory WHERE barrel_no=?", [barrel_no]).fetchone()
        if not row or row["status"] != "in":
            raise HTTPException(404, f"桶号 {barrel_no} 不存在")
        loc = row["location"]
        conn.execute("DELETE FROM barrel_inventory WHERE barrel_no=?", [barrel_no])
        conn.execute("INSERT INTO barrel_log (barrel_no, action, location, created_at) VALUES (?,?,?,?)",
                     [barrel_no, "删除", loc, now])
    return {"success": True, "msg": f"桶号 {barrel_no}（in / {loc}）已删除"}

@app.post("/api/barrel/delete/batch")
def delete_barrel_batch(req: DeleteRequest):
    now = now_str()
    deleted = []
    with get_db() as conn:
        for bn in req.barrel_nos:
            bn = bn.strip()
            if not bn: continue
            row = conn.execute("SELECT * FROM barrel_inventory WHERE barrel_no=?", [bn]).fetchone()
            if row and row["status"] == "in":
                conn.execute("DELETE FROM barrel_inventory WHERE barrel_no=?", [bn])
                conn.execute("INSERT INTO barrel_log (barrel_no, action, location, created_at) VALUES (?,?,?,?)",
                             [bn, "删除", row["location"], now])
                deleted.append(bn)
    return {"success": True, "deleted_count": len(deleted), "deleted_list": deleted}

@app.get("/api/logs")
def get_logs(limit: int = 100):
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM barrel_log ORDER BY id DESC LIMIT ?", [limit]).fetchall()
    return [dict(r) for r in rows]

@app.get("/api/export")
def export_inventory():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM barrel_inventory WHERE status='in' ORDER BY location").fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "桶号库存"
    hf = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center")
    bd = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for col, h in enumerate(["桶号", "储位", "状态", "更新时间"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, bd
    df = Font(name="微软雅黑", size=10)
    da = Alignment(horizontal="center", vertical="center")
    for i, row in enumerate(rows, 2):
        for j, k in enumerate(["barrel_no", "location", "status", "update_time"], 1):
            c = ws.cell(row=i, column=j, value=row[k])
            c.font, c.alignment, c.border = df, da, bd
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 22
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=barrel_inventory.xlsx"})

@app.get("/api/backup")
def trigger_backup():
    """手动触发备份到 Turso"""
    backup_to_turso()
    return {"success": True, "msg": "备份完成"}

# ─── 静态文件 ─────────────────────────────────────────────
app.mount("/", StaticFiles(directory="static", html=True), name="static")

init_db()
print("✅ SQLite 本地数据库就绪")

# 启动时自动备份一次 + 导入 Turso 数据到本地
try:
    turso_data = turso("SELECT * FROM barrel_inventory WHERE status='in'")
    if turso_data:
        with get_db() as conn:
            for row in turso_data:
                conn.execute("INSERT OR REPLACE INTO barrel_inventory (barrel_no, location, status, update_time) VALUES (?,?,?,?)",
                             [row["barrel_no"], row["location"], row["status"], row["update_time"]])
        print(f"✅ 从 Turso 导入 {len(turso_data)} 条数据")
    backup_to_turso()
except Exception as e:
    print(f"⚠️ Turso 同步失败（不影响本地使用）: {e}")
